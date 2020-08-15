from openpyxl import load_workbook

# x = sheet.rows
# list1 = []
# for i in next(x):
#     list1.append(i.value)
# print(list1)
# print(sheet.max_row)
from AUTO_API_TEST.common.pro_path import conf_path
from AUTO_API_TEST.conf.read_config import read_config

class Do_excel:
    def __init__(self,filepath,sheetname):
        self.filepath = filepath
        self.sheetname = sheetname

    def do_excel(self):
        flag = read_config(conf_path,'TESTCASE','flag')
        case_id_list = read_config(conf_path,'TESTCASE','case_id_list')
        wb = load_workbook(self.filepath)
        sheet = wb[self.sheetname]
        test_data=[]
        for i in range(2,sheet.max_row+1):
            dict={}
            dict['case_id']=sheet.cell(i,1).value
            dict['url']=sheet.cell(i,2).value
            dict['param']=sheet.cell(i,3).value
            dict['method']=sheet.cell(i,4).value
            dict['expected']=sheet.cell(i,5).value
            dict['row']=i
            test_data.append(dict)
        if flag == 'on':
            final_list = test_data
        else:
            final_list=[]
            for item in eval(case_id_list):
                final_list.append(test_data[item-1])
        return final_list

    def write(self,row,col,value):
        wb = load_workbook(self.filepath)
        sheet = wb[self.sheetname]
        sheet.cell(row,col).value=value
        wb.save(self.filepath)


# if __name__ == '__main__':
#     response = do_excel(r'E:\workspace\autotest\AUTO_API_TEST\test_data\test_cases.xlsx','test_data')
#     print(response)


